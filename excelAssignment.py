# Names: Kimball Berrett, Matt Vance, Devin Holderness, Parker Francis, Tanner Child, Ethan Smith
# IS303 Section 4 Group 14, 3/31/25
# This program takes unorganized data from an excel file located within the same folder and
# creates a new excel file with the organized data

# import openpyxl and other necessary libraries
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Specify where we are getting the input data to organize
input_file = "Poorly_Organized_Data_1.xlsx"
wb_input = openpyxl.load_workbook(input_file)
sheet_input = wb_input.active

# Upload the data from that file into a list, so that we can work with the list throughout the rest of the program
data = []
for row in sheet_input.iter_rows(min_row=2, values_only=True):
    data.append(row)

# Create a new workbook and remove the automatically created sheet
wb = Workbook()
wb.remove(wb["Sheet"])

# Create the sheets within the workbook based on the subjects, found in the 0 index of our data list
for i in range(0, len(data)):
    if data[i][0] in wb.sheetnames:
        sheet = wb[data[i][0]]
    else:
        sheet = wb.create_sheet(data[i][0])

# Create a list with the column titles so that it is easier to reference in the loop below
column_titles = ['Last Name', 'First Name', 'Student ID', 'Grade']

# Add the column titles to each sheet, and format them with bold, spacing, and add a filter
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    col_num = 1  
    for title in column_titles:
        sheet.cell(row=1, column=col_num, value=title)
        sheet.cell(row=1, column=col_num).font = Font(bold=True)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(title) + 5
        col_num += 1
    max_col = sheet.max_column
    sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}1"

# Populate the new sheets with the data by looping through the list from the input file
for i in range(0, len(data)):
    sheet = wb[data[i][0]]
    next_row = sheet.max_row + 1

    # Use the split function to grab the first, last, and ID from within the same column
    split_string = data[i][1].split("_")
    lastName, firstName, studentID = split_string

    sheet.cell(row=next_row, column=1, value=lastName)
    sheet.cell(row=next_row, column=2, value=firstName)
    sheet.cell(row=next_row, column=3, value=studentID)
    sheet.cell(row=next_row, column=4, value=data[i][2])

# On each sheet, add the summary statistic labels and format them as necessary
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet.cell(row=1,column=6, value="Summary Statistics")
    sheet.cell(row=1,column=6).font = Font(bold=True)
    sheet.cell(row=2,column=6, value="Highest Grade")
    sheet.cell(row=3,column=6, value="Lowest Grade")
    sheet.cell(row=4,column=6, value="Mean Grade")
    sheet.cell(row=5,column=6, value="Median Grade")
    sheet.cell(row=6,column=6, value="Number of Students")
    sheet.cell(row=1,column=7, value="Value")
    sheet.cell(row=1,column=7).font = Font(bold=True)
    sheet.column_dimensions["F"].width = len("Summary Statistics") + 5
    sheet.column_dimensions["G"].width = len("Grade") + 5
    # Create a list to store the grades, and sort it, to access the data and find the median/mean
    grades = []
    for row in sheet.iter_rows(min_col=4, max_col=4, min_row=2, max_row=sheet.max_row):
        for cell in row:
            grades.append(cell.value)
    grades.sort()
    # Add the summary statistics based off the formulas used to calculate them
    sheet.cell(row=2,column=7, value=max(grades))
    sheet.cell(row=3,column=7, value=min(grades))
    sheet.cell(row=4,column=7, value=round((sum(grades)/len(grades)), 1))
    # Calculate the median based off if there is an odd or even number of data points in the list
    if len(grades) % 2 == 1:
        sheet.cell(row=5,column=7, value=grades[len(grades)//2])
    else:
        median = (grades[len(grades)//2] + grades[(len(grades)//2) - 1]) / 2
        sheet.cell(row=5,column=7, value=median)
    sheet.cell(row=6,column=7, value=len(grades))

# Save and export the file
wb.save(filename="formatted_grades.xlsx")